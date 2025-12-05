# backend/converters.py
import os
import subprocess
from pathlib import Path
from typing import Literal
from PIL import Image
from pdf2docx import Converter
import PyPDF2
from openpyxl import Workbook

# Use /app/storage in Docker, or local ./storage for development
STORAGE_DIR = Path(os.environ.get("STORAGE_DIR", "./storage"))
UPLOAD_DIR = STORAGE_DIR / "uploads"
CONVERTED_DIR = STORAGE_DIR / "converted"

UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
CONVERTED_DIR.mkdir(parents=True, exist_ok=True)

# Which conversions we support and how
ConversionKind = Literal["copy", "libreoffice", "pdf_to_image", "image_convert", "pdf_to_docx", "pdf_to_xlsx", "pdf_to_pptx", "image_to_pdf"]

# NOTE: PDF->DOCX/XLSX/PPTX now supported via pdf2docx library
CONVERSION_MAP: dict[tuple[str, str], ConversionKind] = {
    # PDF conversions
    ("pdf", "pdf"): "copy",
    ("pdf", "jpg"): "pdf_to_image",
    ("pdf", "jpeg"): "pdf_to_image",
    ("pdf", "png"): "pdf_to_image",
    ("pdf", "docx"): "pdf_to_docx",
    ("pdf", "doc"): "pdf_to_docx",  # Convert to DOCX then to DOC
    ("pdf", "xlsx"): "pdf_to_xlsx",
    ("pdf", "xls"): "pdf_to_xlsx",  # Convert to XLSX then to XLS
    ("pdf", "pptx"): "pdf_to_pptx",
    ("pdf", "ppt"): "pdf_to_pptx",  # Convert to PPTX then to PPT
    
    # Office to PDF (works with LibreOffice)
    ("doc", "pdf"): "libreoffice",
    ("docx", "pdf"): "libreoffice",
    ("ppt", "pdf"): "libreoffice",
    ("pptx", "pdf"): "libreoffice",
    ("xls", "pdf"): "libreoffice",
    ("xlsx", "pdf"): "libreoffice",
    
    # Office format conversions within same category (Word/Excel/PowerPoint)
    ("doc", "docx"): "libreoffice",
    ("docx", "doc"): "libreoffice",
    ("doc", "doc"): "copy",
    ("docx", "docx"): "copy",
    ("ppt", "pptx"): "libreoffice",
    ("pptx", "ppt"): "libreoffice",
    ("ppt", "ppt"): "copy",
    ("pptx", "pptx"): "copy",
    ("xls", "xlsx"): "libreoffice",
    ("xlsx", "xls"): "libreoffice",
    ("xls", "xls"): "copy",
    ("xlsx", "xlsx"): "copy",
    
    # Cross-office format conversions (Word to Excel, Excel to PowerPoint, etc.)
    ("doc", "xlsx"): "libreoffice",
    ("doc", "xls"): "libreoffice",
    ("doc", "pptx"): "libreoffice",
    ("doc", "ppt"): "libreoffice",
    ("docx", "xlsx"): "libreoffice",
    ("docx", "xls"): "libreoffice",
    ("docx", "pptx"): "libreoffice",
    ("docx", "ppt"): "libreoffice",
    ("xls", "docx"): "libreoffice",
    ("xls", "doc"): "libreoffice",
    ("xls", "pptx"): "libreoffice",
    ("xls", "ppt"): "libreoffice",
    ("xlsx", "docx"): "libreoffice",
    ("xlsx", "doc"): "libreoffice",
    ("xlsx", "pptx"): "libreoffice",
    ("xlsx", "ppt"): "libreoffice",
    ("ppt", "docx"): "libreoffice",
    ("ppt", "doc"): "libreoffice",
    ("ppt", "xlsx"): "libreoffice",
    ("ppt", "xls"): "libreoffice",
    ("pptx", "docx"): "libreoffice",
    ("pptx", "doc"): "libreoffice",
    ("pptx", "xlsx"): "libreoffice",
    ("pptx", "xls"): "libreoffice",
    
    # Office to images (via LibreOffice to PDF, then PDF to image)
    ("doc", "jpg"): "pdf_to_image",
    ("doc", "jpeg"): "pdf_to_image",
    ("doc", "png"): "pdf_to_image",
    ("docx", "jpg"): "pdf_to_image",
    ("docx", "jpeg"): "pdf_to_image",
    ("docx", "png"): "pdf_to_image",
    ("xls", "jpg"): "pdf_to_image",
    ("xls", "jpeg"): "pdf_to_image",
    ("xls", "png"): "pdf_to_image",
    ("xlsx", "jpg"): "pdf_to_image",
    ("xlsx", "jpeg"): "pdf_to_image",
    ("xlsx", "png"): "pdf_to_image",
    ("ppt", "jpg"): "pdf_to_image",
    ("ppt", "jpeg"): "pdf_to_image",
    ("ppt", "png"): "pdf_to_image",
    ("pptx", "jpg"): "pdf_to_image",
    ("pptx", "jpeg"): "pdf_to_image",
    ("pptx", "png"): "pdf_to_image",
    
    # Image format conversions (using Pillow)
    ("jpg", "png"): "image_convert",
    ("jpeg", "png"): "image_convert",
    ("png", "jpg"): "image_convert",
    ("png", "jpeg"): "image_convert",
    ("jpg", "jpg"): "copy",
    ("jpeg", "jpeg"): "copy",
    ("png", "png"): "copy",
    
    # Images to PDF (embed image in PDF)
    ("jpg", "pdf"): "image_to_pdf",
    ("jpeg", "pdf"): "image_to_pdf",
    ("png", "pdf"): "image_to_pdf",
}


def run_libreoffice_convert(input_path: Path, target_ext: str) -> Path:
    """
    Use 'soffice' (LibreOffice) to convert an office document/PDF
    into another office format. This assumes LibreOffice is installed
    in the Docker image.
    """
    # LibreOffice uses output directory + format
    output_dir = CONVERTED_DIR
    cmd = [
        "soffice",
        "--headless",
        "--convert-to",
        target_ext,   # e.g. "docx", "pptx", "xlsx", "pdf"
        "--outdir",
        str(output_dir),
        str(input_path),
    ]
    result = subprocess.run(cmd, capture_output=True, text=True)

    if result.returncode != 0:
        raise RuntimeError(
            f"LibreOffice conversion failed: {result.stderr.strip()}"
        )

    # LibreOffice keeps the same base name but changes extension
    output_path = output_dir / (input_path.stem + f".{target_ext}")
    if not output_path.exists():
        raise FileNotFoundError(
            f"Expected output file not found: {output_path.name}"
        )
    return output_path


def run_pdf_to_images(input_path: Path, target_ext: str) -> Path:
    """
    Convert first page of PDF to a single JPG/PNG using ImageMagick.
    """
    output_path = CONVERTED_DIR / (input_path.stem + f".{target_ext}")
    # ImageMagick command to convert first page only
    cmd = [
        "convert",
        "-density", "300",  # High quality
        f"{input_path}[0]",  # first page
        "-quality", "95",
        str(output_path),
    ]
    result = subprocess.run(cmd, capture_output=True, text=True)
    if result.returncode != 0:
        raise RuntimeError(
            f"Image conversion failed: {result.stderr.strip()}"
        )
    if not output_path.exists():
        raise FileNotFoundError(f"Expected image not found: {output_path.name}")
    return output_path


def run_image_convert(input_path: Path, target_ext: str) -> Path:
    """
    Convert between image formats using Pillow (JPG <-> PNG).
    """
    output_path = CONVERTED_DIR / (input_path.stem + f".{target_ext}")
    try:
        with Image.open(input_path) as img:
            # Convert RGBA to RGB for JPEG
            if target_ext.lower() in ["jpg", "jpeg"] and img.mode == "RGBA":
                rgb_img = Image.new("RGB", img.size, (255, 255, 255))
                rgb_img.paste(img, mask=img.split()[3] if len(img.split()) == 4 else None)
                rgb_img.save(output_path, "JPEG", quality=95)
            else:
                img.save(output_path)
        return output_path
    except Exception as e:
        raise RuntimeError(f"Image conversion failed: {str(e)}")


def run_pdf_to_docx(input_path: Path) -> Path:
    """
    Convert PDF to DOCX using pdf2docx library.
    """
    output_path = CONVERTED_DIR / (input_path.stem + ".docx")
    try:
        cv = Converter(str(input_path))
        cv.convert(str(output_path))
        cv.close()
        
        if not output_path.exists():
            raise FileNotFoundError(f"Expected DOCX file not found: {output_path.name}")
        return output_path
    except Exception as e:
        raise RuntimeError(f"PDF to DOCX conversion failed: {str(e)}")


def run_pdf_to_xlsx(input_path: Path) -> Path:
    """
    Convert PDF to XLSX by extracting text and organizing it into rows.
    This is a basic conversion - extracts text line by line.
    """
    output_path = CONVERTED_DIR / (input_path.stem + ".xlsx")
    try:
        # Extract text from PDF
        with open(input_path, 'rb') as file:
            pdf_reader = PyPDF2.PdfReader(file)
            
            # Create Excel workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Extracted Data"
            
            current_row = 1
            for page_num, page in enumerate(pdf_reader.pages, 1):
                text = page.extract_text()
                lines = text.split('\n')
                
                # Add page header
                ws.cell(row=current_row, column=1, value=f"--- Page {page_num} ---")
                current_row += 1
                
                # Add each line as a row
                for line in lines:
                    if line.strip():  # Skip empty lines
                        ws.cell(row=current_row, column=1, value=line.strip())
                        current_row += 1
                
                # Add blank row between pages
                current_row += 1
            
            wb.save(str(output_path))
        
        if not output_path.exists():
            raise FileNotFoundError(f"Expected XLSX file not found: {output_path.name}")
        return output_path
    except Exception as e:
        raise RuntimeError(f"PDF to XLSX conversion failed: {str(e)}")


def run_pdf_to_pptx(input_path: Path) -> Path:
    """
    Convert PDF to PPTX by embedding each page as an image in slides.
    """
    output_path = CONVERTED_DIR / (input_path.stem + ".pptx")
    try:
        # Use LibreOffice to convert PDF to PPTX (it will embed pages as images)
        return run_libreoffice_convert(input_path, "pptx")
    except Exception as e:
        raise RuntimeError(f"PDF to PPTX conversion failed: {str(e)}")


def run_image_to_pdf(input_path: Path) -> Path:
    """
    Convert image (JPG/PNG) to PDF by embedding it in a PDF page.
    """
    output_path = CONVERTED_DIR / (input_path.stem + ".pdf")
    try:
        with Image.open(input_path) as img:
            # Convert to RGB if needed (PDF doesn't support transparency well)
            if img.mode == "RGBA":
                rgb_img = Image.new("RGB", img.size, (255, 255, 255))
                rgb_img.paste(img, mask=img.split()[3] if len(img.split()) == 4 else None)
                rgb_img.save(output_path, "PDF", resolution=100.0)
            else:
                img.save(output_path, "PDF", resolution=100.0)
        
        if not output_path.exists():
            raise FileNotFoundError(f"Expected PDF file not found: {output_path.name}")
        return output_path
    except Exception as e:
        raise RuntimeError(f"Image to PDF conversion failed: {str(e)}")


def perform_conversion(
    input_path: Path,
    target_ext: str,
) -> Path:
    """
    Core conversion dispatcher. Figures out which converter to call
    based on (src_ext, target_ext) using CONVERSION_MAP.
    """
    src_ext = input_path.suffix.lower().lstrip(".")  # "pdf"
    target_ext = target_ext.lower().lstrip(".")

    key = (src_ext, target_ext)
    if key not in CONVERSION_MAP:
        raise ValueError(
            f"Conversion from .{src_ext} to .{target_ext} is not supported."
        )

    kind = CONVERSION_MAP[key]

    if kind == "copy":
        # e.g. PDF -> PDF: just copy
        output_path = CONVERTED_DIR / (input_path.stem + f".{target_ext}")
        output_path.write_bytes(input_path.read_bytes())
        return output_path

    if kind == "libreoffice":
        return run_libreoffice_convert(input_path, target_ext)

    if kind == "pdf_to_image":
        return run_pdf_to_images(input_path, target_ext)

    if kind == "image_convert":
        return run_image_convert(input_path, target_ext)

    if kind == "pdf_to_docx":
        return run_pdf_to_docx(input_path)

    if kind == "pdf_to_xlsx":
        return run_pdf_to_xlsx(input_path)

    if kind == "pdf_to_pptx":
        return run_pdf_to_pptx(input_path)

    if kind == "image_to_pdf":
        return run_image_to_pdf(input_path)

    # Fallback (should never hit)
    raise RuntimeError("Unknown conversion kind encountered.")


# ============================================================================
# PDF MERGE, SPLIT, and COMPRESS
# ============================================================================

def merge_pdfs(file_paths: list[Path]) -> Path:
    """Merge multiple PDF files into one."""
    if not file_paths:
        raise ValueError("No PDF files provided for merging.")
    
    if len(file_paths) < 2:
        raise ValueError("At least 2 PDF files are required for merging.")
    
    # Verify all files are PDFs
    for file_path in file_paths:
        if file_path.suffix.lower() != ".pdf":
            raise ValueError(f"File {file_path.name} is not a PDF.")
    
    # Create merged PDF
    merger = PyPDF2.PdfMerger()
    
    try:
        for file_path in file_paths:
            merger.append(str(file_path))
        
        # Save merged PDF
        output_path = CONVERTED_DIR / f"merged_{file_paths[0].stem}.pdf"
        with open(output_path, "wb") as output_file:
            merger.write(output_file)
        
        merger.close()
        return output_path
    
    except Exception as e:
        merger.close()
        raise RuntimeError(f"Failed to merge PDFs: {str(e)}")


def split_pdf(file_path: Path, page_ranges: str) -> list[Path]:
    """
    Split a PDF into multiple files based on page ranges.
    page_ranges format: "1-3,5,7-9" (pages are 1-indexed)
    Returns a list of paths to the split PDF files.
    """
    if file_path.suffix.lower() != ".pdf":
        raise ValueError("File must be a PDF.")
    
    try:
        reader = PyPDF2.PdfReader(str(file_path))
        total_pages = len(reader.pages)
        
        # Parse page ranges
        ranges = []
        for part in page_ranges.split(","):
            part = part.strip()
            if "-" in part:
                start, end = part.split("-")
                start, end = int(start), int(end)
                if start < 1 or end > total_pages or start > end:
                    raise ValueError(f"Invalid page range: {part}")
                ranges.append((start - 1, end))  # Convert to 0-indexed
            else:
                page = int(part)
                if page < 1 or page > total_pages:
                    raise ValueError(f"Invalid page number: {page}")
                ranges.append((page - 1, page))  # Convert to 0-indexed
        
        # Create split PDFs
        output_paths = []
        for i, (start, end) in enumerate(ranges):
            writer = PyPDF2.PdfWriter()
            for page_num in range(start, end):
                writer.add_page(reader.pages[page_num])
            
            output_path = CONVERTED_DIR / f"{file_path.stem}_part{i+1}.pdf"
            with open(output_path, "wb") as output_file:
                writer.write(output_file)
            
            output_paths.append(output_path)
        
        return output_paths
    
    except ValueError as e:
        raise ValueError(str(e))
    except Exception as e:
        raise RuntimeError(f"Failed to split PDF: {str(e)}")


def compress_pdf(file_path: Path, quality: str = "medium") -> Path:
    """
    Compress a PDF using Ghostscript.
    quality: "low", "medium", "high" (lower quality = smaller file size)
    """
    if file_path.suffix.lower() != ".pdf":
        raise ValueError("File must be a PDF.")
    
    # Ghostscript quality settings
    quality_map = {
        "low": "/screen",      # 72 dpi, smallest file
        "medium": "/ebook",    # 150 dpi, balanced
        "high": "/printer",    # 300 dpi, best quality
    }
    
    if quality not in quality_map:
        quality = "medium"
    
    gs_quality = quality_map[quality]
    output_path = CONVERTED_DIR / f"{file_path.stem}_compressed.pdf"
    
    # Ghostscript command
    gs_cmd = [
        "gs",
        "-sDEVICE=pdfwrite",
        "-dCompatibilityLevel=1.4",
        f"-dPDFSETTINGS={gs_quality}",
        "-dNOPAUSE",
        "-dQUIET",
        "-dBATCH",
        f"-sOutputFile={output_path}",
        str(file_path),
    ]
    
    try:
        result = subprocess.run(gs_cmd, check=True, capture_output=True, text=True)
        
        if not output_path.exists():
            raise RuntimeError("Ghostscript did not produce output file.")
        
        return output_path
    
    except subprocess.CalledProcessError as e:
        raise RuntimeError(f"Ghostscript failed: {e.stderr}")
    except Exception as e:
        raise RuntimeError(f"Failed to compress PDF: {str(e)}")

